import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from collections import Counter

import analysis as ay

##FETCHING##
filePath = "C:/Users/emagr/Documents/School/Y3S2/FYP/FYP Statistics.xlsx"
dataset = pd.read_excel(filePath, sheet_name="Anxiety")
workbook = load_workbook(filePath)

##PROCESSING##
def processing(row):
    tokenList = []
    tSentList = []
    emojiList = []
    eSentList = []
    histogramList = []
    faceList = []
    objList = []

    #Used for frequency calculations
    allTokens = []

    for name, group in dataset.groupby('Name'):
        allTokens.clear

        for index, row in dataset.iterrows():
            
            print(f"\nProcessing row {index+1}/{len(dataset)}")
            caption = row['Caption']
            image = cv2.imread(row['Image Reference'])

            ##CAPTIONS##
            words = ay.tokenize(caption)
            cleaned, emojis = ay.clean(words)
            emojiSenti = ay.emojiSentiment(emojis)
            tokenSenti = ay.tokenSentiment(cleaned)

            allTokens.append(cleaned)

            ##IMAGES##
            colour = ay.colourExtraction(image)
            face = ay.facialExtraction(image)
            obj = ay.objectExtraction(image)

            ##APPENDING##
            tokenList.append(cleaned)
            tSentList.append(tokenSenti)
            emojiList.append(emojis)
            eSentList.append(emojiSenti)
            histogramList.append(colour)
            objList.append(obj)
            faceList.append(face)

            ##DEBUGGING##
            print(f"Tokens: {words}")
            print(f"Clean: {cleaned}")
            print(f"Emojis: {emojis}")
            print(f"Emoji Sentiment: {emojiSenti}")
            print(f"Caption Sentiment: {tokenSenti}")
            print(f"Colour: {colour}")
            print(f"Emotions: {face}")
            print(f"Objects: {obj}")
        
        ##FREQUENCIES##
        tFreq = ay.tokenFrequency(allTokens)
        nFreq = ay.nGramFrequency(allTokens)

        ##DEBUGGING##
        print(f"Token Freq {tFreq}")
        print(f"NGrams Freq: {nFreq}")
    
    ##DATAFRAME CREATION##
    toAdd = pd.DataFrame({
        'Tokens': tokenList,
        'Token Sentiment Score': tSentList,
        'Emojis': emojiList,
        'Emoji Sentiment Score': eSentList,
        'Image Colour Histogram': histogramList,
        'Emotions Detected in Image': faceList,
        'Objects Detected in Image': objList
    })

    return toAdd


##SAVING##
def saveToExcel(toAdd):
    with pd.ExcelWriter(filePath, engine='openpyxl', mode = 'a', if_sheet_exists='overlay') as writer:
        writer.book = workbook
        writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        toAdd.to_excel(writer, sheet_name = "Anxiety", index = False, header = False, startrow = writer.sheets['Anxiety'].max_row)

##RUNNING##
if __name__ == "__main__":
    newData = processing(dataset)
    #saveToExcel(newData)