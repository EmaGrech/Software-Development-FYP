import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from collections import Counter

import analysis as ay

##FETCHING##
filePath = "C:/Users/emagr/Documents/School/Y3S2/FYP/FYP Statistics.xlsx"
sheetName = "Anxiety"
dataset = pd.read_excel(filePath, sheet_name = sheetName)

##PROCESSING##
def processing(row):
    tokenList = []
    tSentList = []
    tFreqList = []
    nFreqList = []
    emojiList = []
    eSentList = []
    histogramList = []
    faceList = []
    objList = []

    #Used for frequency calculations
    allTokens = []

    for name, group in dataset.groupby('Name'):
        allTokens = []

        for index, row in group.iterrows():
            
            print(f"\nProcessing row {index+1}/{len(dataset)}")
            caption = row['Caption']
            imageRef = row['Image Reference']
            image = cv2.imread(imageRef)

            ##CAPTIONS##
            words = ay.tokenize(caption)
            cleaned, emojis = ay.clean(words)
            emojiSenti = ay.emojiSentiment(emojis)
            tokenSenti = ay.tokenSentiment(cleaned)

            allTokens.append(cleaned)

            ##IMAGES##
            colour = ay.colourExtraction(image)
            face = ay.facialExtraction(image)
            obj = ay.objectExtraction(image)

            ##APPENDING##
            tokenList.append(cleaned)
            tSentList.append(tokenSenti)
            emojiList.append(emojis)
            eSentList.append(emojiSenti)
            histogramList.append(colour)
            objList.append(obj)
            faceList.append(face)

            ##DEBUGGING##
            #print(f"Tokens: {words}")
            #print(f"Clean: {cleaned}")
            #print(f"Emojis: {emojis}")
            #print(f"Emoji Sentiment: {emojiSenti}")
            #print(f"Caption Sentiment: {tokenSenti}")
            #print(f"Colour: {colour}")
            #print(f"Emotions: {face}")
            #print(f"Objects: {obj}")
        
        ##FREQUENCIES##
        tFreq = ay.tokenFrequency(allTokens)
        nFreq = ay.nGramFrequency(allTokens)
        #Converting to String for cv
        tFreqConvert = ', '.join([f'{k}: {v}' for k, v in tFreq.items()])
        nFreqConvert = ', '.join([f'{k}: {v}' for k, v in nFreq.items()])

        tFreqList.append(tFreqConvert)
        nFreqList.append(nFreqConvert)

        ##DEBUGGING##
        #print(f"Token Freq {tFreq}")
        #print(f"NGrams Freq: {nFreq}")
    
    ##DATAFRAME CREATION##
    toAdd = pd.DataFrame({
        'Tokens': tokenList,
        'Token Sentiment Score': tSentList,
        'Emojis': emojiList,
        'Emoji Sentiment Score': eSentList,
        'Image Colour Histogram': histogramList,
        'Emotions Detected in Image': faceList,
        'Objects Detected in Image': objList
    })

    toAdd2 = pd.DataFrame({
        'Name': dataset['Name'].unique(),
        'Token Frequency': tFreqList,
        'n-Gram Frequency': nFreqList
    })

    return toAdd, toAdd2


##SAVING##
def saveToExcel(toAdd, toAdd2):
    #Saving data that does not require merging
    with pd.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        sheet = writer.sheets[sheetName]
        startrow = 1
        startcol = sheet.max_column  
        toAdd.to_excel(writer, sheet_name=sheetName, index=False, header=False, startrow=startrow, startcol=startcol)
    book = load_workbook(filePath)
    sheet = book[sheetName]

    col1 = 'K'  
    col2 = 'L' 

    curr = sheet.max_row - len(toAdd) + 1

    #Saving data that does requires merging
    for name, group in dataset.groupby('Name'):
        cells = len(group)
        start = curr
        end = start + cells - 1
    
        tFreq = toAdd2[toAdd2['Name'] == name]['Token Frequency'].values[0]
        nFreq = toAdd2[toAdd2['Name'] == name]['n-Gram Frequency'].values[0]
        sheet.merge_cells(f'{col1}{start}:{col1}{end}')
        sheet[f'{col1}{start}'] = tFreq
        sheet.merge_cells(f'{col2}{start}:{col2}{end}')
        sheet[f'{col2}{start}'] = nFreq

        curr = end + 1
    
    book.save(filePath)
    

##RUNNING##
if __name__ == "__main__":
    newData, newData2 = processing(dataset)
    saveToExcel(newData, newData2)