#Saving the feature extraction

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from collections import Counter

import analysis as ay

##PROCESSING##
def processing(dataset):
    tokenList = []
    tSentList = []
    tFreqList = []
    nFreqList = []
    emojiList = []
    eSentList = []
    histogramList = []
    faceList = []
    objList = []

    allTokens = []

    for name, group in dataset.groupby('Name'):
        individualsTokens = []

        for index, row in group.iterrows():
            
            print(f"\nProcessing row {index+1}/{len(dataset)}")
            caption = row['Caption']
            imageRef = row['Image Reference']
            image = cv2.imread(imageRef)

            ##CAPTIONS##
            words = ay.tokenize(caption)
            cleaned, emojis = ay.clean(words)
            emojiSenti = ay.emojiSentiment(emojis)
            tokenSenti = ay.tokenSentiment(cleaned)

            individualsTokens.extend(cleaned)

            ##IMAGES##
            colour = ay.colourExtraction(image)
            face = ay.facialExtraction(image)
            obj = ay.objectExtraction(image)

            ##APPENDING##
            tokenList.append(cleaned)
            tSentList.append(tokenSenti)
            emojiList.append(emojis)
            eSentList.append(emojiSenti)
            histogramList.append(colour)
            objList.append(obj)
            faceList.append(face)
        
        allTokens.append(individualsTokens)
        
    ##FREQUENCIES##
    overallTokens = [token for person in allTokens for token in person]
    topTokens = ay.tokenFrequency([overallTokens])  
    topNGrams = ay.nGramFrequency([overallTokens])  
    
    topTokenSet = set(topTokens.keys())
    topNGramSet = set(topNGrams.keys())

    tFreqList = []
    nFreqList = []

    #Cross-checking everyone's top frequencies with eachother's tokens
    for individualsTokens in allTokens:
        tFreq = Counter(token for token in individualsTokens if token in topTokenSet)
        tFreq = {token: count / len(individualsTokens) for token, count in tFreq.items()}

        nGrams = [' '.join(individualsTokens[i:i+2]) for i in range(len(individualsTokens) - 2 + 1)]
        nFreq = Counter(ngram for ngram in nGrams if ngram in topNGramSet)
        
        totalNGrams = len(nGrams)
        nFreq = {ngram: count / totalNGrams for ngram, count in nFreq.items()}
        
        #Converting to string for cv
        tFreqConvert = ', '.join([f'{k}: {v:.4f}' for k, v in tFreq.items()])
        nFreqConvert = ', '.join([f'{k}: {v:.4f}' for k, v in nFreq.items()])

        tFreqList.append(tFreqConvert)
        nFreqList.append(nFreqConvert)
    
    ##DATAFRAME CREATION##
    toAdd = pd.DataFrame({
        'Tokens': tokenList,
        'Token Sentiment Score': tSentList,
        'Emojis': emojiList,
        'Emoji Sentiment Score': eSentList,
        'Image Colour Histogram': histogramList,
        'Emotions Detected in Image': faceList,
        'Objects Detected in Image': objList
    })

    toAdd2 = pd.DataFrame({
        'Name': dataset['Name'].unique(),
        'Token Frequency': tFreqList,
        'n-Gram Frequency': nFreqList
    })

    return toAdd, toAdd2


##SAVING##
def saveToExcel(toAdd, toAdd2, filePath, sheetName, dataset):
    #Saving data that does not require merging
    with pd.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        sheet = writer.sheets[sheetName]
        startrow = 1
        startcol = sheet.max_column  
        toAdd.to_excel(writer, sheet_name = sheetName, index = False, header = False, startrow = startrow, startcol = startcol)
    book = load_workbook(filePath)
    sheet = book[sheetName]

    existing = [cell.value for cell in sheet[1]]

    cols = toAdd.columns.tolist() 
    cols2 = toAdd2.columns.tolist()

    start = {}
    firstRow = 2

    for col in cols:
        if col in existing:
            colID = existing.index(col) + 1  
            letter = chr(ord('A') + colID - 1)  
            start[col] = letter

    #Saving data that does requires merging
    for name, group in dataset.groupby('Name'):
        for col in cols2:
            if col in existing:
                colID = existing.index(col) + 1
                letter = chr(ord('A') + colID - 1)
                data = toAdd2[toAdd2['Name'] == name][col].values[0]

                curr = firstRow + list(dataset['Name']).index(name)
                
                sheet[f'{letter}{curr}'] = data

    book.save(filePath)
    

##RUNNING##
if __name__ == "__main__":
    filePath = 'C:/Users/emagr/Documents/School/Y3S2/FYP/FYP Statistics.xlsx'
    xls = pd.ExcelFile(filePath)
    sheetNames = xls.sheet_names

    for sheetName in sheetNames:
        dataset = pd.read_excel(filePath, sheet_name = sheetName)

        newData, newData2 = processing(dataset)
        saveToExcel(newData, newData2, filePath, sheetName, dataset)